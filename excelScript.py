import openpyxl
import openpyxl.styles
import database

class Table4:
    # Метод для форматирования границ таблицы
    def set_border(ws, cell_range, need_to_medium, need_to_medium_up, need_to_medium_down):
        thin = openpyxl.styles.Side(border_style="thin", color="000000")
        medium = openpyxl.styles.Side(border_style="medium", color="000000")
        for row in ws[cell_range]:
            for cell in row:
                if cell == row[len(row) - 1]:
                    cell.border = openpyxl.styles.Border(top=thin, left=thin, right=medium, bottom=thin)
                    cell.fill = openpyxl.styles.PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                else:
                    cell.border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
            if len(need_to_medium) != 0:
                if row == ws[need_to_medium[0]]:
                    for cell in row:
                        cell.border = openpyxl.styles.Border(top=medium, left=medium, right=medium, bottom=medium)
                    need_to_medium.pop(0)
        
            if len(need_to_medium_up) != 0:
                if row == ws[need_to_medium_up[0]]:
                    for cell in row:
                        if cell == row[len(row) - 1]:
                            cell.border = openpyxl.styles.Border(top=medium, left=thin, right=medium, bottom=thin)
                        else:
                            cell.border = openpyxl.styles.Border(top=medium, left=thin, right=thin, bottom=thin)
                    need_to_medium_up.pop(0)
            if len(need_to_medium_down) != 0:
                if row == ws[need_to_medium_down[0]]:
                    for cell in row:
                        if cell == row[len(row) - 1]:
                            cell.border = openpyxl.styles.Border(top=thin, left=thin, right=medium, bottom=medium)
                        else:
                            cell.border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=medium)

                    need_to_medium_down.pop(0)
            for row in ws[cell_range]:
                for cell in row:
                    if cell == row[0]:
                        cell.border = openpyxl.styles.Border(top=medium, left=medium, right=thin, bottom=medium)
                        
                        


    # Метод для формирования Excel таблицы
    def make_excel():
        # Получаем метрики из базы данных
        data = database.Database.get_metrics()
        # Создаем книгу и заполняем ее
        wb = openpyxl.load_workbook("./template.xltx")
        wb.template = False

        # получаем лист, с которым будем работать
        sheet = wb['Метрики']
        ws = wb.active

        sections = database.Database.get_sections()
        current_row = 5
        current_category = 0
        counter = 0
        need_to_medium = []
        need_to_medium_up = [3]
        need_to_medium_down = []
        for row in range(len(data)):

            # Вставка категорий
            if current_category != int(data[row].section_id):
                ws.insert_rows(current_row)
                current_category += 1
                sheet.cell(row=current_row, column=1).value = sections[current_category - 1].description
                sheet.cell(row=current_row, column=1).font = openpyxl.styles.Font(size=4)
                ws.row_dimensions[current_row].height = 10
                ws.merge_cells(start_column=1, start_row=current_row, end_column=11, end_row=current_row)
                need_to_medium.append(str(current_row))
                current_row += 1

            formatted_data = data[row].to_array()
            # Заполнение строк
            for col in range(len(formatted_data) - 1):
                cell = sheet.cell(row=current_row, column=col + 1)
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='center', vertical='center')

                # Пропуска записи в недоступную ячейку после слияния
                try:
                    cell.value = formatted_data[col]
                except Exception as err:
                    continue

            # Слияние ячеек с номерам критериев
            if row + 1 < len(data):
                if formatted_data[0] == data[row+1].to_array()[0]:
                    counter += 1
                    if counter == 1:
                        need_to_medium_up.append(current_row)
                else:
                    if counter == 0 and formatted_data[1] is None:
                        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
                    else:
                        ws.merge_cells(start_row=current_row - counter, start_column=1, end_row=current_row, end_column=1)
                        counter = 0
                        need_to_medium_down.append(current_row)
            else:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
                need_to_medium_down.append(current_row)

            current_row += 1

        Table4.set_border(ws, 'A3:K' + str(current_row - 1), need_to_medium, need_to_medium_up, need_to_medium_down)

        # Сохранение файла
        wb.save('Метрики.xlsx')
        print("Файл успешно сохранен")


Table4.make_excel()
